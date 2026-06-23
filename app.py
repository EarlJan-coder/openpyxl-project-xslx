import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename) # load the worksheet
    sheet = wb['Sheet1'] # Getting reference from first Sheet


    for row in range(2, sheet.max_row + 1): # Iterate every row in the xlsx and stores it in **row** variable.
        cell = sheet.cell(row, 3) # this locates the position of cell: row = the iterated value( from range 2 to max row + 1 which means row 2 to row 4), 3 = the column number
        corrected_price = cell.value * 0.9 # this calculates the cell targeted at **cell**
        corrected_price_cell = sheet.cell(row, 4) # adds new column
        corrected_price_cell.value = corrected_price # assign values to the column
        
    values = Reference(sheet, 
            min_row= 2, # specifies the minimum row number from which data should be extracted. In this case, it's set to 2
            max_row= sheet.max_row, # sets the maximum row number for extraction. Here, sheet.max_row is used, which retrieves the last row number in the sheet. This ensures that all rows are extracted from the current range.
            min_col= 4, # This specifies the minimum column number from which data should be extracted. It's set to 4.
            max_col= 4 # This sets the maximum column number for extraction
            )

    chart = BarChart()  # This initializes a new instance of the BarChart class, which is used to create a bar chart.
    chart.add_data(values)  # The add_data method adds data to the chart. Here, 'values' is a reference object that points to the range of cells in the worksheet where the values for the chart are located (columns 4 and above from row 2 to the last row).
    sheet.add_chart(chart, 'e2')  # The add_chart method adds the created chart to the worksheet. 'e2' specifies the cell where the chart should be placed.

    wb.save(filename) #saves into a new file